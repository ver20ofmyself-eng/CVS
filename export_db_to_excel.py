#!/usr/bin/env python3
"""
Экспорт всех таблиц БД CVS Analyzer в Excel.

Создаёт папку  export_YYYY-MM-DD_HH-MM/  рядом со скриптом
и сохраняет каждую таблицу в отдельный .xlsx файл,
а также сводный файл  ALL_TABLES.xlsx  с отдельными листами.

Установка зависимостей (один раз):
    pip install psycopg2-binary openpyxl pandas

Запуск:
    python export_db_to_excel.py

С другой БД:
    DATABASE_URL=postgresql://... python export_db_to_excel.py
"""

import os
import sys
from datetime import datetime
from pathlib import Path

# ── Зависимости ────────────────────────────────────────────────────────────────
try:
    import psycopg2
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Установите зависимости:\n   pip install psycopg2-binary openpyxl pandas")
    sys.exit(1)

# ── Конфигурация ───────────────────────────────────────────────────────────────
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://cvs_user:cvs_password@localhost:5433/cvs_analyzer"
)

# Таблицы в порядке выгрузки
TABLES = [
    ("users",         "Пользователи"),
    ("tariffs",       "Тарифы"),
    ("user_tariffs",  "Тарифы пользователей"),
    ("vacancies",     "Вакансии"),
    ("analyses",      "Анализы"),
    ("payments",      "Платежи"),
    ("prompts",       "Промпты"),
    ("prompt_history","История промптов"),
]

# ── Стили ──────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="106AB7")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
EVEN_FILL = PatternFill("solid", fgColor="EAE2D7")


def style_sheet(ws):
    """Применяет стили к листу."""
    max_col = ws.max_column
    max_row = ws.max_row

    # Заголовок
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Данные
    for row in range(2, max_row + 1):
        fill = EVEN_FILL if row % 2 == 0 else None
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

    # Авторазмер колонок
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 200)):   # сканируем первые 200 строк
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 4, 12)

    # Фиксируем шапку
    ws.freeze_panes = "A2"


def export():
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_dir = Path(__file__).parent / f"export_{timestamp}"
    out_dir.mkdir(exist_ok=True)

    print(f"📁 Папка экспорта: {out_dir}")

    # Подключение к БД
    try:
        conn = psycopg2.connect(DATABASE_URL)
        print("✅ Подключение к БД установлено.\n")
    except Exception as e:
        print(f"❌ Ошибка подключения к БД: {e}")
        print("   Убедитесь что Docker запущен и база доступна.")
        sys.exit(1)

    all_frames = {}

    for table_name, table_label in TABLES:
        print(f"  ⏳ Экспорт: {table_label} ({table_name})...", end="", flush=True)
        try:
            df = pd.read_sql(f"SELECT * FROM {table_name} ORDER BY id", conn)

            if df.empty:
                print(f" пусто (0 строк)")
                all_frames[table_label] = df
                continue

            # Маскируем чувствительные поля
            if "password_hash" in df.columns:
                df["password_hash"] = "***скрыто***"
            if "api_key_encrypted" in df.columns:
                df["api_key_encrypted"] = "***скрыто***"

            # Форматируем даты
            for col in df.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M")
            for col in df.select_dtypes(include=["datetime64[ns]"]).columns:
                df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M")

            # JSON поля → строки
            for col in df.columns:
                if df[col].dtype == object:
                    df[col] = df[col].apply(
                        lambda x: str(x) if isinstance(x, (dict, list)) else x
                    )

            # Сохраняем отдельный файл
            file_path = out_dir / f"{table_name}.xlsx"
            df.to_excel(file_path, index=False, sheet_name=table_label[:31])
            wb = load_workbook(file_path)
            style_sheet(wb.active)
            wb.save(file_path)

            all_frames[table_label] = df
            print(f" ✅ {len(df)} строк → {file_path.name}")

        except Exception as e:
            print(f" ⚠️  ошибка: {e}")
            all_frames[table_label] = pd.DataFrame()

    # Сводный файл ALL_TABLES.xlsx
    print("\n  📊 Создание сводного файла ALL_TABLES.xlsx...")
    all_path = out_dir / "ALL_TABLES.xlsx"
    with pd.ExcelWriter(all_path, engine="openpyxl") as writer:
        for label, df in all_frames.items():
            sheet_name = label[:31]
            if df.empty:
                pd.DataFrame({"(нет данных)": []}).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Применяем стили к каждому листу
    wb = load_workbook(all_path)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(all_path)

    conn.close()

    print(f"""
╔══════════════════════════════════════════════════════╗
║          ✅ Экспорт завершён успешно!                ║
╠══════════════════════════════════════════════════════╣
║  Папка:        export_{timestamp:<30}║
║  Отдельные:    по одному файлу на каждую таблицу     ║
║  Сводный:      ALL_TABLES.xlsx (все листы вместе)    ║
╚══════════════════════════════════════════════════════╝
    """)


if __name__ == "__main__":
    export()
